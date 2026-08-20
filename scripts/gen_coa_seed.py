# One-time generator: builds seed/coa.json (canonical Monarch chart of accounts)
# from the real Yardi upload CSV (canonical upload order) and the FHND budget
# workbook (names + display order). Re-run only if the COA changes.
import csv, json, re, sys
from openpyxl import load_workbook

CSV_PATH = r"C:\Users\TroySteiss\Downloads\FHND Budget Revision 06062026.csv"
CSV_PATH2 = r"C:\Users\TroySteiss\Downloads\PHND Revised Budget 06062026.csv"
XLSX_PATH = r"C:\Users\TroySteiss\OneDrive - MIMG\FHND Budget 04222026 (RMC) (002).xlsx"
OUT = r"C:\Users\TroySteiss\nd-budget-tool\seed\coa.json"

def csv_codes(path):
    codes = []
    with open(path, newline="") as f:
        rows = list(csv.reader(f))
    in_detail = False
    for r in rows:
        if r and r[0].startswith("//BudgetDetail:"):
            in_detail = True
            continue
        if in_detail and len(r) >= 2 and r[1].strip():
            codes.append(r[1].strip())
    return codes

codes_f = csv_codes(CSV_PATH)
codes_p = csv_codes(CSV_PATH2)
assert codes_f == codes_p, "CSV account order differs between FHND and PHND!"
csv_order = {c: i + 1 for i, c in enumerate(codes_f)}
print(f"CSV detail accounts: {len(codes_f)}")

wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
ws = wb["FHNDBudget"]
names = {}          # code -> name
display = {}        # code -> budget-sheet row (display order)
for row in ws.iter_rows(min_row=1, max_row=466, min_col=1, max_col=2):
    a, b = row[0].value, row[1].value
    if a is None:
        continue
    code = str(a).strip()
    if not re.fullmatch(r"\d{3,4}", code):
        continue
    name = str(b).strip() if b else ""
    if code not in names:
        names[code] = name
        display[code] = row[0].row
print(f"Budget-sheet coded rows: {len(names)}")

# ---- section / kind / pcode rules (from the reverse-engineered structure) ----
# sections: (section key, header codes, total codes, pcode default, detail predicate)
def n(c): return int(c)

TOTALS = {"5004","5029","5049","5070","5190","5500","6170","6370","6399","6470",
          "6570","6670","6770","6870","6970","7070","7098","7099","7279","7280",
          "7315","7500","8200","8602","8950","9000"}
HEADERS = {"4991","4992","4993","5017","5030","5100","5999","6000","6100","6300",
           "6371","6373","6400","6500","6600","6690","6700","6800","6900","7000",
           "7290","7320","8499","8900"}
NAME_FIXES = {"7491": "SP INTEREST INCOME"}

UTIL_INCOME = {"5122","5164","5167","5169","5170","5171","5172","5174"}  # reimbursements -> pcode 4

def classify(code):
    c = n(code)
    if code in TOTALS: return ("total", None)
    if code in HEADERS: return ("header", None)
    if c in (3080, 3090, 3091):        return ("detail", None)      # principal (below the line)
    if c in (4994, 4995, 4996):        return ("detail", "1")
    if c == 5003:                      return ("detail", "loss")
    if 5018 <= c <= 5028:              return ("detail", "2")
    if 5031 <= c <= 5049:              return ("detail", "3")
    if 5100 <= c <= 5189:
        return ("detail", "4" if code in UTIL_INCOME else "5")
    if c in (6108, 6110):              return ("detail", "6")
    if c == 6112:                      return ("detail", "7")
    if c in (6113, 6116, 6117):       return ("detail", "8")
    if 6100 <= c <= 6170:              return ("detail", "9")   # remaining fixed admin
    if 6300 <= c <= 6392:              return ("detail", "9")
    if 6400 <= c <= 6469:              return ("detail", "10")
    if 6500 <= c <= 6569:              return ("detail", "11")
    if 6600 <= c <= 6669:              return ("detail", "12")
    if 6700 <= c <= 6999:              return ("detail", "13")
    if 7000 <= c <= 7069:              return ("detail", "14")
    if 7300 <= c <= 7302:              return ("detail", None)       # interest
    if 7321 <= c <= 7499:              return ("detail", None)       # special projects
    if c >= 8200:                      return ("detail", None)       # below NOI
    return ("detail", None)

def section_of(code):
    c = n(code)
    if c in (3080, 3090, 3091): return "principal"
    if c <= 5004: return "gpr"
    if c <= 5029: return "concessions"
    if c <= 5049 or code == "5070": return "rental_loss"
    if c <= 5500: return "other_income"
    if c <= 6170: return "fixed_admin"
    if c <= 6370: return "admin"
    if c <= 6399 or (6373 <= c <= 6392): return "corporate_events"
    if c <= 6470: return "payroll"
    if c <= 6570: return "marketing"
    if c <= 6670: return "utilities"
    if c <= 6770: return "in_house_maint"
    if c <= 6870: return "exterior_cam"
    if c <= 6970: return "contract_services"
    if c <= 7070: return "rehab_replacement"
    if c <= 7280: return "totals"
    if c <= 7315: return "interest"
    if c <= 7500: return "special_projects"
    if c <= 8602: return "below_noi"
    return "below_noi"

# seasonal curve defaults (monthly weight names resolved in app)
CURVES = {"6818": "snow", "6934": "snow",
          "6604": "electric", "6605": "electric", "6606": "electric", "6607": "electric",
          "6608": "heat", "6609": "heat", "6610": "heat", "6611": "heat",
          "6724": "summer", "6924": "summer"}

# corporate events header/total codes seen: 6373 header? total 6399. Ensure 6399/6373 handled:
all_codes = sorted(set(list(csv_order.keys()) + list(names.keys())), key=lambda c: (n(c) if n(c) >= 3080 and n(c) < 3200 else n(c)))
# order: use budget-sheet display row when known, else large offset by numeric code
out = []
for code in all_codes:
    kind, pcode = classify(code)
    out.append({
        "code": code,
        "name": names.get(code) or NAME_FIXES.get(code, ""),
        "kind": kind,
        "section": section_of(code),
        "pcode": pcode,
        "csv_order": csv_order.get(code),
        "display_order": display.get(code, 100000 + n(code)),
        "curve": CURVES.get(code, "flat") if kind == "detail" else None,
    })
out.sort(key=lambda r: r["display_order"])

missing_names = [r["code"] for r in out if not r["name"]]
not_in_csv = [r["code"] for r in out if r["csv_order"] is None and r["kind"] == "detail"]
totals_in_csv = [c for c in csv_order if c in TOTALS or c in HEADERS]
print("codes missing names:", missing_names)
print("detail codes NOT in csv:", not_in_csv)
print("header/total codes present in csv:", totals_in_csv)
print("total records:", len(out))

with open(OUT, "w") as f:
    json.dump(out, f, indent=1)
print("wrote", OUT)
